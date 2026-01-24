"""
Vistas para exportación de reportes a Excel
"""
import os
import decimal
import openpyxl
from decimal import Decimal
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Q
from http import HTTPStatus

from farm import settings
from .models import CashFlow
from ..hrm.models import Subsidiary


@csrf_exempt
def export_sales_report_excel(request):
    """Exportar reporte de ventas a Excel"""
    if request.method == 'POST':
        try:
            # Obtener datos del reporte
            report_date = request.POST.get('report_date')
            subsidiary_id = request.POST.get('subsidiary')
            cash_id = request.POST.get('cash_account')
            subsidiary_obj = None
            
            if not report_date:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una fecha'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Filtrar cashflows del día por sucursal
            if subsidiary_id and subsidiary_id != '0':
                subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary_id))
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date,
                    cash__subsidiary_id=subsidiary_id
                )
            else:
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date
                )
            
            cashflows = cashflows.select_related('cash', 'user', 'cash__subsidiary', 'order', 'order__client', 'order__subsidiary').prefetch_related('order__orderdetail_set')
            
            # Filtrar cashflows con order_id (ventas) y sin order_id (gastos)
            order_cashflows = cashflows.filter(order__isnull=False, order__status__in=['P', 'C'])
            
            # Obtener todas las órdenes del día del reporte
            orders_of_day = Order.objects.filter(
                register_date=report_date,
                status__in=['P', 'C']
            ).order_by('id')
            
            if subsidiary_id and subsidiary_id != '0':
                orders_of_day = orders_of_day.filter(subsidiary_id=subsidiary_id)
            
            orders_of_day = orders_of_day.select_related('client', 'subsidiary', 'user').prefetch_related('orderdetail_set')
            
            # ========================================
            # NUEVA ESTRUCTURA: ADELANTOS DEL DÍA (sin pagos totales)
            # ========================================
            advances_of_day = {}
            
            for order in orders_of_day:
                order_advances = cashflows.filter(
                    order=order,
                    type='E',
                    order_type_entry='A',
                    transaction_date=report_date
                ).order_by('id')
                
                if order_advances.exists():
                    # Verificar si la orden tiene pagos totales del día
                    order_total_payments = cashflows.filter(
                        order=order,
                        type='E',
                        order_type_entry='T',
                        transaction_date=report_date
                    )
                    
                    # Solo incluir en adelantos si NO tiene pagos totales del día
                    if not order_total_payments.exists():
                        total_advances = sum(float(cf.total) for cf in order_advances)
                        saldo = float(order.total) - total_advances
                        
                        advances_of_day[f"advance_{order.id}"] = {
                            'order': order,
                            'cashflows': list(order_advances),
                            'total_amount': total_advances,
                            'balance': saldo,
                            'cashflow_count': order_advances.count()
                        }
            
            # ========================================
            # NUEVA ESTRUCTURA: PAGOS TOTALES DEL DÍA
            # ========================================
            full_payments_of_day = {}
            
            for order in orders_of_day:
                order_payments = cashflows.filter(
                    order=order,
                    type='E',
                    order_type_entry='T',
                    transaction_date=report_date
                ).order_by('id')
                
                if order_payments.exists():
                    # También incluir adelantos del día si existen
                    order_advances = cashflows.filter(
                        order=order,
                        type='E',
                        order_type_entry='A',
                        transaction_date=report_date
                    ).order_by('id')
                    
                    # Combinar adelantos y pagos totales
                    all_cashflows = list(order_advances) + list(order_payments)
                    total_payments = sum(float(cf.total) for cf in all_cashflows)
                    
                    full_payments_of_day[f"payment_{order.id}"] = {
                        'order': order,
                        'cashflows': all_cashflows,
                        'total_amount': total_payments,
                        'cashflow_count': len(all_cashflows)
                    }
            
            # ========================================
            # COMBINAR EN day_income
            # ========================================
            day_income = {}
            day_income.update(advances_of_day)
            day_income.update(full_payments_of_day)
            
            # ========================================
            # PAGOS DE FECHAS ANTERIORES
            # ========================================
            previous_payments_cashflows = cashflows.filter(
                type='E',
                order_type_entry='T',
                transaction_date=report_date,
                order__register_date__lt=report_date
            ).order_by('order_id', 'id')
            
            # Preparar datos de egresos
            expenses_cashflows = cashflows.filter(
                order__isnull=True,
                type='S'
            )
            
            # Calcular totales de apertura
            total_apertura = cashflows.filter(type='A').aggregate(total=Sum('total'))['total'] or 0
            
            # Calcular totales de ingresos del día
            total_day_income = 0
            day_income_cash = 0
            day_income_yape = 0
            day_income_deposit = 0
            
            for key, data in day_income.items():
                total_day_income += data['total_amount']
                for cashflow in data['cashflows']:
                    if cashflow.way_to_pay == 'E':
                        day_income_cash += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'Y':
                        day_income_yape += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'D':
                        day_income_deposit += decimal.Decimal(cashflow.total)
            
            # Calcular totales de pagos anteriores
            total_previous_payments = 0
            previous_payments_cash = 0
            previous_payments_yape = 0
            previous_payments_deposit = 0
            
            for cashflow in previous_payments_cashflows:
                total_previous_payments += float(cashflow.total)
                if cashflow.way_to_pay == 'E':
                    previous_payments_cash += decimal.Decimal(cashflow.total)
                elif cashflow.way_to_pay == 'Y':
                    previous_payments_yape += decimal.Decimal(cashflow.total)
                elif cashflow.way_to_pay == 'D':
                    previous_payments_deposit += decimal.Decimal(cashflow.total)
            
            total_expenses_amount = expenses_cashflows.aggregate(total=Sum('total'))['total'] or 0
            
            total_cash = day_income_cash + previous_payments_cash
            total_yape = day_income_yape + previous_payments_yape
            total_deposit = day_income_deposit + previous_payments_deposit
            total_general = total_cash + total_yape + total_deposit + total_apertura
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Reporte {report_date}"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill_primary = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
            header_fill_success = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            header_fill_danger = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            title_font = Font(bold=True, size=16, color="007bff")
            border = Border(
                left=Side(style='medium', color='adb5bd'),
                right=Side(style='medium', color='adb5bd'),
                top=Side(style='medium', color='adb5bd'),
                bottom=Side(style='medium', color='adb5bd')
            )
            # Formato de contabilidad con S/
            currency_format = '_("S/"* #,##0.00_);_("S/"* (#,##0.00);_("S/"* "-"??_);_(@_)'
            
            # Título principal
            ws.merge_cells('A1:J1')
            ws['A1'] = f"TIENDA: {subsidiary_obj.name.upper() if subsidiary_obj else 'TODAS'} - DÍA: {datetime.strptime(report_date, '%Y-%m-%d').strftime('%d-%m-%Y')}"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Sección de INGRESOS DEL DÍA
            ws['A3'] = "INGRESOS DEL DÍA"
            ws['A3'].font = header_font
            ws['A3'].fill = header_fill_primary
            ws.merge_cells('A3:J3')
            ws['A3'].alignment = Alignment(horizontal='center')
            
            # Encabezados de ingresos
            income_headers = ['N° CPTE.', 'CLIENTE O RAZON SOCIAL', 'CANT.', 'DESCRIPCIÓN DEL PRODUCTO', 'USUARIO', 'TIPO PAGO', 'A CUENTA S/.', 'SALDO S/.', 'TOTAL S/.']
            for col, header in enumerate(income_headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_primary
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de ingresos del día
            row = 5
            
            # Primero: Adelantos (sin pagos totales)
            for key, data in advances_of_day.items():
                for i, cashflow in enumerate(data['cashflows']):
                    if i == 0:  # Primera fila con datos de la orden
                        ws.cell(row=row, column=1, value=f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}").border = border
                        ws.cell(row=row, column=2, value=data['order'].client.full_name if data['order'].client else '-').border = border
                        ws.cell(row=row, column=3, value=1).border = border  # Cantidad
                        # Descripción del producto
                        product_desc = ""
                        if data['order'].orderdetail_set.exists():
                            product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                        else:
                            product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                        ws.cell(row=row, column=4, value=product_desc).border = border
                    else:
                        # Filas adicionales sin datos de orden
                        ws.cell(row=row, column=1, value="").border = border
                        ws.cell(row=row, column=2, value="").border = border
                        ws.cell(row=row, column=3, value="").border = border
                        ws.cell(row=row, column=4, value="").border = border
                    
                    # Datos del cashflow
                    ws.cell(row=row, column=5, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                    
                    # Tipo de pago
                    payment_type = ""
                    if cashflow.way_to_pay == 'E':
                        payment_type = "EFECTIVO"
                    elif cashflow.way_to_pay == 'Y':
                        payment_type = "YAPE"
                    elif cashflow.way_to_pay == 'D':
                        payment_type = "DEPÓSITO"
                    ws.cell(row=row, column=6, value=payment_type).border = border
                    
                    cell_total = ws.cell(row=row, column=7, value=Decimal(cashflow.total))
                    cell_total.border = border
                    cell_total.number_format = currency_format
                    
                    if i == 0:  # Solo en la primera fila
                        cell_saldo = ws.cell(row=row, column=8, value=Decimal(data['balance']))
                        cell_saldo.border = border
                        cell_saldo.number_format = currency_format
                        
                        cell_order_total = ws.cell(row=row, column=9, value=Decimal(data['order'].total))
                        cell_order_total.border = border
                        cell_order_total.number_format = currency_format
                    else:
                        ws.cell(row=row, column=8, value="").border = border
                        ws.cell(row=row, column=9, value="").border = border
                    
                    row += 1
            
            # Separador visual (fila vacía con color)
            if advances_of_day and full_payments_of_day:
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col, value="")
                    cell.fill = PatternFill(start_color="bbdefb", end_color="bbdefb", fill_type="solid")
                    cell.border = border
                row += 1
            
            # Segundo: Pagos totales del día (con fondo celeste)
            for key, data in full_payments_of_day.items():
                for i, cashflow in enumerate(data['cashflows']):
                    if i == 0:  # Primera fila con datos de la orden
                        cell1 = ws.cell(row=row, column=1, value=f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}")
                        cell1.border = border
                        cell1.fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
                        
                        cell2 = ws.cell(row=row, column=2, value=data['order'].client.full_name if data['order'].client else '-')
                        cell2.border = border
                        cell2.fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
                        
                        cell3 = ws.cell(row=row, column=3, value=1)
                        cell3.border = border
                        cell3.fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
                        
                        # Descripción del producto
                        product_desc = ""
                        if data['order'].orderdetail_set.exists():
                            product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                        else:
                            product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                        cell4 = ws.cell(row=row, column=4, value=product_desc)
                        cell4.border = border
                        cell4.fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
                    else:
                        # Filas adicionales sin datos de orden
                        for col in range(1, 5):
                            cell = ws.cell(row=row, column=col, value="")
                            cell.border = border
                            cell.fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
                    
                    # Datos del cashflow
                    cell5 = ws.cell(row=row, column=5, value=cashflow.user.first_name or cashflow.user.username or '-')
                    cell5.border = border
                    cell5.fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
                    
                    # Tipo de pago
                    payment_type = ""
                    if cashflow.way_to_pay == 'E':
                        payment_type = "EFECTIVO"
                    elif cashflow.way_to_pay == 'Y':
                        payment_type = "YAPE"
                    elif cashflow.way_to_pay == 'D':
                        payment_type = "DEPÓSITO"
                    cell6 = ws.cell(row=row, column=6, value=payment_type)
                    cell6.border = border
                    cell6.fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
                    
                    cell7 = ws.cell(row=row, column=7, value=decimal.Decimal(cashflow.total))
                    cell7.border = border
                    cell7.fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
                    cell7.number_format = currency_format
                    
                    if i == 0:
                        cell8 = ws.cell(row=row, column=8, value="PAGADO")
                        cell8.border = border
                        cell8.fill = PatternFill(start_color="e8f5e8", end_color="e8f5e8", fill_type="solid")
                        cell8.font = Font(bold=True, color="2e7d32")
                        
                        cell9 = ws.cell(row=row, column=9, value=Decimal(data['order'].total))
                        cell9.border = border
                        cell9.fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
                        cell9.number_format = currency_format
                    else:
                        cell8 = ws.cell(row=row, column=8, value="")
                        cell8.border = border
                        cell8.fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
                        
                        cell9 = ws.cell(row=row, column=9, value="")
                        cell9.border = border
                        cell9.fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
                    
                    row += 1
            
            # Totales de ingresos
            row += 1
            ws.cell(row=row, column=8, value="YAPE:").font = Font(bold=True)
            cell_yape = ws.cell(row=row, column=9, value=decimal.Decimal(day_income_yape))
            cell_yape.font = Font(bold=True)
            cell_yape.number_format = currency_format
            row += 1
            ws.cell(row=row, column=8, value="EFECTIVO:").font = Font(bold=True)
            cell_efectivo = ws.cell(row=row, column=9, value=decimal.Decimal(day_income_cash))
            cell_efectivo.font = Font(bold=True)
            cell_efectivo.number_format = currency_format
            row += 1
            ws.cell(row=row, column=8, value="TOTAL INGRESOS:").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=8).fill = header_fill_primary
            cell_total_ing = ws.cell(row=row, column=9, value=decimal.Decimal(total_day_income))
            cell_total_ing.font = Font(bold=True, color="FFFFFF")
            cell_total_ing.fill = header_fill_primary
            cell_total_ing.number_format = currency_format
            
            # Sección de SALDOS (Pagos de fechas anteriores)
            row += 3
            ws.cell(row=row, column=1, value="SALDOS")
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=1).fill = header_fill_success
            ws.merge_cells(f'A{row}:F{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Encabezados de saldos
            row += 1
            saldos_headers = ['N° CPTE.', 'FECHA', 'DESCRIPCIÓN', 'USUARIO', 'TIPO PAGO', 'S/TOTAL']
            for col, header in enumerate(saldos_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_success
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de saldos (pagos de fechas anteriores)
            row += 1
            for cashflow in previous_payments_cashflows:
                ws.cell(row=row, column=1, value=f"{cashflow.order.subsidiary.serial}-{cashflow.order.correlative:03d}").border = border
                ws.cell(row=row, column=2, value=cashflow.order.register_date.strftime('%d-%m-%Y')).border = border
                
                # Descripción con productos
                desc = cashflow.description or "PAGO TOTAL"
                desc += f" (Usuario: {cashflow.order.user.first_name or cashflow.order.user.username or '-'})"
                if cashflow.order.orderdetail_set.exists():
                    products = " | ".join([detail.product_name or "Producto Manual" for detail in cashflow.order.orderdetail_set.all()])
                    desc += f" - {products}"
                ws.cell(row=row, column=3, value=desc).border = border
                
                ws.cell(row=row, column=4, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                
                payment_type = ""
                if cashflow.way_to_pay == 'E':
                    payment_type = "EFECTIVO"
                elif cashflow.way_to_pay == 'Y':
                    payment_type = "YAPE"
                elif cashflow.way_to_pay == 'D':
                    payment_type = "DEPÓSITO"
                ws.cell(row=row, column=5, value=payment_type).border = border
                cell_saldo_total = ws.cell(row=row, column=6, value=decimal.Decimal(cashflow.total))
                cell_saldo_total.border = border
                cell_saldo_total.number_format = currency_format
                row += 1
            
            # Totales de saldos
            row += 1
            ws.cell(row=row, column=5, value="YAPE:").font = Font(bold=True)
            cell_saldo_yape = ws.cell(row=row, column=6, value=decimal.Decimal(previous_payments_yape))
            cell_saldo_yape.font = Font(bold=True)
            cell_saldo_yape.number_format = currency_format
            row += 1
            ws.cell(row=row, column=5, value="EFECTIVO:").font = Font(bold=True)
            cell_saldo_efectivo = ws.cell(row=row, column=6, value=decimal.Decimal(previous_payments_cash))
            cell_saldo_efectivo.font = Font(bold=True)
            cell_saldo_efectivo.number_format = currency_format
            row += 1
            ws.cell(row=row, column=5, value="TOTAL CANCELACIONES:").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=5).fill = header_fill_success
            cell_total_cancel = ws.cell(row=row, column=6, value=decimal.Decimal(total_previous_payments))
            cell_total_cancel.font = Font(bold=True, color="FFFFFF")
            cell_total_cancel.fill = header_fill_success
            cell_total_cancel.number_format = currency_format
            
            # Sección de EGRESOS
            row += 3
            ws.cell(row=row, column=1, value="EGRESOS")
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=1).fill = header_fill_danger
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Encabezados de egresos
            row += 1
            egresos_headers = ['NRO', 'DESCRIPCIÓN', 'TIPO EGRESO', 'USUARIO', 'MONTO']
            for col, header in enumerate(egresos_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_danger
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de egresos
            row += 1
            for i, cashflow in enumerate(expenses_cashflows, 1):
                ws.cell(row=row, column=1, value=i).border = border
                ws.cell(row=row, column=2, value=cashflow.description or '-').border = border
                
                expense_type = ""
                if cashflow.type_expense == 'V':
                    expense_type = "VARIABLE"
                elif cashflow.type_expense == 'F':
                    expense_type = "FIJO"
                elif cashflow.type_expense == 'P':
                    expense_type = "PERSONAL"
                elif cashflow.type_expense == 'O':
                    expense_type = "OTRO"
                ws.cell(row=row, column=3, value=expense_type).border = border
                ws.cell(row=row, column=4, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                cell_egreso = ws.cell(row=row, column=5, value=decimal.Decimal(cashflow.total))
                cell_egreso.border = border
                cell_egreso.number_format = currency_format
                row += 1
            
            # Total de egresos
            row += 1
            ws.cell(row=row, column=4, value="TOTAL EGRESOS:").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=4).fill = header_fill_danger
            cell_total_egreso = ws.cell(row=row, column=5, value=decimal.Decimal(total_expenses_amount))
            cell_total_egreso.font = Font(bold=True, color="FFFFFF")
            cell_total_egreso.fill = header_fill_danger
            cell_total_egreso.number_format = currency_format
            
            # Sección de RESUMENES
            row += 3
            ws.cell(row=row, column=1, value="RESUMENES")
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=1).fill = header_fill_success
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Resumen de ingresos
            row += 1
            ws.cell(row=row, column=1, value="INGRESOS").font = Font(bold=True, size=12, color="007bff")
            row += 1
            ws.cell(row=row, column=1, value="APERTURA DE CAJA:").font = Font(bold=True)
            cell_res_apertura = ws.cell(row=row, column=2, value=decimal.Decimal(total_apertura))
            cell_res_apertura.font = Font(bold=True)
            cell_res_apertura.number_format = currency_format
            row += 1
            ws.cell(row=row, column=1, value="INGRESOS DEL DÍA:").font = Font(bold=True)
            cell_res_ing_dia = ws.cell(row=row, column=2, value=decimal.Decimal(total_day_income))
            cell_res_ing_dia.font = Font(bold=True)
            cell_res_ing_dia.number_format = currency_format
            row += 1
            ws.cell(row=row, column=1, value="SALDOS:").font = Font(bold=True)
            cell_res_saldos = ws.cell(row=row, column=2, value=decimal.Decimal(total_previous_payments))
            cell_res_saldos.font = Font(bold=True)
            cell_res_saldos.number_format = currency_format
            row += 1
            ws.cell(row=row, column=1, value="SUBTOTAL INGRESOS:").font = Font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
            cell_res_subtotal = ws.cell(row=row, column=2, value=decimal.Decimal(total_apertura + total_day_income + total_previous_payments))
            cell_res_subtotal.font = Font(bold=True)
            cell_res_subtotal.fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
            cell_res_subtotal.number_format = currency_format
            
            # Resumen de egresos
            row += 2
            ws.cell(row=row, column=1, value="EGRESOS").font = Font(bold=True, size=12, color="dc3545")
            row += 1
            ws.cell(row=row, column=1, value="TOTAL EGRESOS:").font = Font(bold=True)
            cell_res_egresos = ws.cell(row=row, column=2, value=decimal.Decimal(total_expenses_amount))
            cell_res_egresos.font = Font(bold=True)
            cell_res_egresos.number_format = currency_format
            
            # Resumen final
            row += 3
            ws.cell(row=row, column=1, value="TOTAL EFECTIVO:").font = Font(bold=True, size=11, color="28a745")
            cell_res_efectivo = ws.cell(row=row, column=2, value=decimal.Decimal(total_cash))
            cell_res_efectivo.font = Font(bold=True, size=11, color="28a745")
            cell_res_efectivo.number_format = currency_format
            row += 1
            ws.cell(row=row, column=1, value="TOTAL YAPE:").font = Font(bold=True, size=11, color="17a2b8")
            cell_res_yape = ws.cell(row=row, column=2, value=decimal.Decimal(total_yape))
            cell_res_yape.font = Font(bold=True, size=11, color="17a2b8")
            cell_res_yape.number_format = currency_format
            row += 1
            ws.cell(row=row, column=1, value="APERTURA CAJA:").font = Font(bold=True, size=11, color="007bff")
            cell_res_apertura2 = ws.cell(row=row, column=2, value=decimal.Decimal(total_apertura))
            cell_res_apertura2.font = Font(bold=True, size=11, color="007bff")
            cell_res_apertura2.number_format = currency_format
            row += 1
            ws.cell(row=row, column=1, value="TOTAL EGRESOS:").font = Font(bold=True, size=11, color="dc3545")
            cell_res_egresos2 = ws.cell(row=row, column=2, value=decimal.Decimal(total_expenses_amount))
            cell_res_egresos2.font = Font(bold=True, size=11, color="dc3545")
            cell_res_egresos2.number_format = currency_format
            row += 1
            ws.cell(row=row, column=1, value="TOTAL FINAL:").font = Font(bold=True, size=12, color="ffc107")
            ws.cell(row=row, column=1).fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            cell_res_final = ws.cell(row=row, column=2, value=decimal.Decimal(total_general - total_expenses_amount))
            cell_res_final.font = Font(bold=True, size=12, color="ffc107")
            cell_res_final.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            cell_res_final.number_format = currency_format
            
            # Ajustar ancho de columnas
            column_widths = [15, 25, 8, 30, 15, 12, 12, 12, 12, 12, 12, 12, 12]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Guardar archivo
            filename = f"reporte_ventas_gastos_{report_date}.xlsx"
            filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            wb.save(filepath)
            
            # Retornar URL del archivo
            file_url = f"{settings.MEDIA_URL}reports/{filename}"
            
            return JsonResponse({
                'success': True,
                'message': 'Reporte exportado exitosamente',
                'file_url': file_url,
                'filename': filename
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al exportar el reporte: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)


@csrf_exempt
def export_subscriptions_report_excel(request):
    """Exportar reporte de ventas por usuario a Excel"""
    if request.method == 'POST':
        try:
            from apps.users.models import CustomUser
            
            # Obtener datos del reporte
            report_date = request.POST.get('report_date')
            user_id = int(request.POST.get('user'))
            user_obj = None
            
            if not report_date:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una fecha'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Filtrar cashflows del día por usuario
            if user_id and user_id != 0:
                user_obj = CustomUser.objects.get(id=user_id)
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date,
                    user_id=user_id
                )
            else:
                cashflows = CashFlow.objects.filter(
                    transaction_date=report_date
                )
            
            cashflows = cashflows.select_related('cash', 'user', 'cash__subsidiary', 'order', 'order__client', 'order__subsidiary').prefetch_related('order__orderdetail_set')
            
            # Filtrar cashflows con order_id (ventas) y sin order_id (gastos)
            order_cashflows = cashflows.filter(order__isnull=False, order__status__in=['P', 'C'])
            
            # Obtener todas las órdenes del día del reporte
            orders_of_day = Order.objects.filter(
                register_date=report_date,
                status__in=['P', 'C']
            ).order_by('id')
            
            if user_id and user_id != 0:
                # Obtener IDs de órdenes que tienen cashflows del usuario del día
                orders_with_user_cashflows = order_cashflows.filter(
                    user_id=user_id
                ).values_list('order_id', flat=True).distinct()
                
                # Filtrar órdenes: creadas por el usuario O con cashflows del usuario
                orders_of_day = orders_of_day.filter(
                    Q(user_id=user_id) | Q(id__in=orders_with_user_cashflows)
                )
            
            orders_of_day = orders_of_day.select_related('client', 'subsidiary', 'user').prefetch_related('orderdetail_set')
            
            # ========================================
            # DICT 1: ADELANTOS DEL USUARIO
            # ========================================
            adelantos_usuario = {}
            
            for order in orders_of_day:
                order_advances = cashflows.filter(
                    order=order,
                    type='E',
                    order_type_entry='A',
                    transaction_date=report_date,
                    user_id=user_id
                ).order_by('id')
                
                if order_advances.exists():
                    # Verificar si la orden tiene pagos totales
                    order_total_payments = cashflows.filter(
                        order=order,
                        type='E',
                        order_type_entry='T',
                        transaction_date=report_date,
                        user_id=user_id
                    )
                    
                    # Solo incluir en adelantos si NO tiene pagos totales
                    if not order_total_payments.exists():
                        total_advances = sum(float(cf.total) for cf in order_advances)
                        saldo = float(order.total) - total_advances
                        
                        adelantos_usuario[f"advance_{order.id}"] = {
                            'tipo': 'adelanto',
                            'order': order,
                            'cashflows': list(order_advances),
                            'total_amount': total_advances,
                            'saldo': saldo,
                            'cashflow_count': order_advances.count()
                        }
            
            # ========================================
            # DICT 2: PAGOS TOTALES DE ÓRDENES DEL USUARIO
            # ========================================
            pagos_totales_usuario = {}
            
            for order in orders_of_day:
                # Solo procesar órdenes creadas por el usuario
                if order.user_id == int(user_id):
                    order_payments = cashflows.filter(
                        order=order,
                        type='E',
                        order_type_entry='T',
                        transaction_date=report_date,
                        user_id=user_id
                    ).order_by('id')
                    
                    if order_payments.exists():
                        # También incluir adelantos si existen
                        order_advances = cashflows.filter(
                            order=order,
                            type='E',
                            order_type_entry='A',
                            transaction_date=report_date,
                            user_id=user_id
                        ).order_by('id')
                        
                        # Combinar adelantos y pagos totales
                        all_cashflows = list(order_advances) + list(order_payments)
                        total_payments = sum(float(cf.total) for cf in all_cashflows)
                        
                        pagos_totales_usuario[f"payment_{order.id}"] = {
                            'tipo': 'pago_total_usuario',
                            'order': order,
                            'cashflows': all_cashflows,
                            'total_amount': total_payments,
                            'cashflow_count': len(all_cashflows)
                        }
            
            # ========================================
            # DICT 3: PAGOS TOTALES DE ÓRDENES NO DEL USUARIO (Cancelaciones)
            # ========================================
            pagos_totales_otros = {}
            
            other_orders_payments = cashflows.filter(
                type='E',
                order_type_entry='T',
                transaction_date=report_date,
                user_id=user_id,
                order__register_date=report_date
            ).exclude(
                order__user_id=user_id
            ).order_by('order_id', 'id')
            
            for cashflow in other_orders_payments:
                pagos_totales_otros[f"cancellation_{cashflow.id}"] = {
                    'tipo': 'pago_total_otros',
                    'order': cashflow.order,
                    'cashflows': [cashflow],
                    'total_amount': float(cashflow.total),
                    'cashflow_count': 1
                }
            
            # ========================================
            # COMBINAR EN ingresos_del_dia
            # ========================================
            ingresos_del_dia = {}
            ingresos_del_dia.update(adelantos_usuario)
            ingresos_del_dia.update(pagos_totales_usuario)
            ingresos_del_dia.update(pagos_totales_otros)
            
            # ========================================
            # PAGOS DE FECHAS ANTERIORES
            # ========================================
            pagos_fechas_anteriores = cashflows.filter(
                type='E',
                order_type_entry='T',
                user_id=user_id,
                transaction_date=report_date,
                order__register_date__lt=report_date
            ).order_by('order_id', 'id')
            
            # Preparar datos de egresos
            expenses_cashflows = cashflows.filter(
                order__isnull=True,
                type='S'
            )
            
            # Calcular totales de apertura
            total_apertura = cashflows.filter(type='A').aggregate(total=Sum('total'))['total'] or 0
            
            # Calcular totales de ingresos del día
            total_ingresos_dia = 0
            ingresos_efectivo = 0
            ingresos_yape = 0
            ingresos_deposito = 0
            
            for key, data in ingresos_del_dia.items():
                total_ingresos_dia += data['total_amount']
                for cashflow in data['cashflows']:
                    if cashflow.way_to_pay == 'E':
                        ingresos_efectivo += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'Y':
                        ingresos_yape += decimal.Decimal(cashflow.total)
                    elif cashflow.way_to_pay == 'D':
                        ingresos_deposito += decimal.Decimal(cashflow.total)
            
            # Calcular totales de pagos anteriores
            total_pagos_anteriores = 0
            pagos_anteriores_efectivo = 0
            pagos_anteriores_yape = 0
            pagos_anteriores_deposito = 0
            
            for cashflow in pagos_fechas_anteriores:
                total_pagos_anteriores += float(cashflow.total)
                if cashflow.way_to_pay == 'E':
                    pagos_anteriores_efectivo += decimal.Decimal(cashflow.total)
                elif cashflow.way_to_pay == 'Y':
                    pagos_anteriores_yape += decimal.Decimal(cashflow.total)
                elif cashflow.way_to_pay == 'D':
                    pagos_anteriores_deposito += decimal.Decimal(cashflow.total)
            
            total_expenses_amount = expenses_cashflows.aggregate(total=Sum('total'))['total'] or 0
            
            total_efectivo = ingresos_efectivo + pagos_anteriores_efectivo
            total_yape = ingresos_yape + pagos_anteriores_yape
            total_deposito = ingresos_deposito + pagos_anteriores_deposito
            total_general = total_efectivo + total_yape + total_deposito + total_apertura
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Reporte {report_date}"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill_primary = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
            header_fill_success = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            header_fill_danger = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            title_font = Font(bold=True, size=16, color="007bff")
            border = Border(
                left=Side(style='medium', color='adb5bd'),
                right=Side(style='medium', color='adb5bd'),
                top=Side(style='medium', color='adb5bd'),
                bottom=Side(style='medium', color='adb5bd')
            )
            # Formato de contabilidad con S/
            currency_format = '_("S/"* #,##0.00_);_("S/"* (#,##0.00);_("S/"* "-"??_);_(@_)'
            
            # Título principal
            ws.merge_cells('A1:J1')
            user_name = f"{user_obj.first_name} {user_obj.last_name}".strip() if user_obj else "TODOS"
            ws['A1'] = f"USUARIO: {user_name.upper()} - DÍA: {datetime.strptime(report_date, '%Y-%m-%d').strftime('%d-%m-%Y')}"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Sección de INGRESOS DEL DÍA
            ws['A3'] = "INGRESOS DEL DÍA"
            ws['A3'].font = header_font
            ws['A3'].fill = header_fill_primary
            ws.merge_cells('A3:J3')
            ws['A3'].alignment = Alignment(horizontal='center')
            
            # Encabezados de ingresos
            income_headers = ['N° CPTE.', 'CLIENTE O RAZON SOCIAL', 'CANT.', 'DESCRIPCIÓN DEL PRODUCTO', 'USUARIO', 'TIPO PAGO', 'A CUENTA S/.', 'SALDO S/.', 'TOTAL S/.']
            for col, header in enumerate(income_headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_primary
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de ingresos del día
            row = 5
            
            # 1. ADELANTOS
            for key, data in adelantos_usuario.items():
                for i, cashflow in enumerate(data['cashflows']):
                    if i == 0:  # Primera fila con datos de la orden
                        ws.cell(row=row, column=1, value=f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}").border = border
                        ws.cell(row=row, column=2, value=data['order'].client.full_name if data['order'].client else '-').border = border
                        ws.cell(row=row, column=3, value=1).border = border
                        product_desc = ""
                        if data['order'].orderdetail_set.exists():
                            product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                        else:
                            product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                        ws.cell(row=row, column=4, value=product_desc).border = border
                    else:
                        for col in range(1, 5):
                            ws.cell(row=row, column=col, value="").border = border
                    
                    ws.cell(row=row, column=5, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                    
                    payment_type = "EFECTIVO" if cashflow.way_to_pay == 'E' else ("YAPE" if cashflow.way_to_pay == 'Y' else "DEPÓSITO")
                    ws.cell(row=row, column=6, value=payment_type).border = border
                    
                    cell_total = ws.cell(row=row, column=7, value=Decimal(cashflow.total))
                    cell_total.border = border
                    cell_total.number_format = currency_format
                    
                    if i == 0:
                        cell_saldo = ws.cell(row=row, column=8, value=Decimal(data['saldo']))
                        cell_saldo.border = border
                        cell_saldo.number_format = currency_format
                        
                        cell_order_total = ws.cell(row=row, column=9, value=Decimal(data['order'].total))
                        cell_order_total.border = border
                        cell_order_total.number_format = currency_format
                    else:
                        ws.cell(row=row, column=8, value="").border = border
                        ws.cell(row=row, column=9, value="").border = border
                    
                    row += 1
            
            # Separador azul
            if adelantos_usuario and (pagos_totales_usuario or pagos_totales_otros):
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col, value="")
                    cell.fill = PatternFill(start_color="bbdefb", end_color="bbdefb", fill_type="solid")
                    cell.border = border
                row += 1
            
            # 2. PAGOS TOTALES DEL USUARIO (fondo verde)
            for key, data in pagos_totales_usuario.items():
                for i, cashflow in enumerate(data['cashflows']):
                    fill_color = PatternFill(start_color="e8f5e8", end_color="e8f5e8", fill_type="solid")
                    
                    if i == 0:
                        cell1 = ws.cell(row=row, column=1, value=f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}")
                        cell1.border = border
                        cell1.fill = fill_color
                        
                        cell2 = ws.cell(row=row, column=2, value=data['order'].client.full_name if data['order'].client else '-')
                        cell2.border = border
                        cell2.fill = fill_color
                        
                        cell3 = ws.cell(row=row, column=3, value=1)
                        cell3.border = border
                        cell3.fill = fill_color
                        
                        product_desc = ""
                        if data['order'].orderdetail_set.exists():
                            product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                        else:
                            product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                        cell4 = ws.cell(row=row, column=4, value=product_desc)
                        cell4.border = border
                        cell4.fill = fill_color
                    else:
                        for col in range(1, 5):
                            cell = ws.cell(row=row, column=col, value="")
                            cell.border = border
                            cell.fill = fill_color
                    
                    cell5 = ws.cell(row=row, column=5, value=cashflow.user.first_name or cashflow.user.username or '-')
                    cell5.border = border
                    cell5.fill = fill_color
                    
                    payment_type = "EFECTIVO" if cashflow.way_to_pay == 'E' else ("YAPE" if cashflow.way_to_pay == 'Y' else "DEPÓSITO")
                    cell6 = ws.cell(row=row, column=6, value=payment_type)
                    cell6.border = border
                    cell6.fill = fill_color
                    
                    cell7 = ws.cell(row=row, column=7, value=decimal.Decimal(cashflow.total))
                    cell7.border = border
                    cell7.fill = fill_color
                    cell7.number_format = currency_format
                    
                    if i == 0:
                        cell8 = ws.cell(row=row, column=8, value="PAGO TOTAL")
                        cell8.border = border
                        cell8.fill = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
                        cell8.font = Font(bold=True, color="1b5e20")
                        
                        cell9 = ws.cell(row=row, column=9, value=Decimal(data['order'].total))
                        cell9.border = border
                        cell9.fill = fill_color
                        cell9.number_format = currency_format
                    else:
                        for col in [8, 9]:
                            cell = ws.cell(row=row, column=col, value="")
                            cell.border = border
                            cell.fill = fill_color
                    
                    row += 1
            
            # Separador naranja
            if pagos_totales_otros:
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col, value="")
                    cell.fill = PatternFill(start_color="ff9800", end_color="ff9800", fill_type="solid")
                    cell.border = border
                row += 1
            
            # 3. CANCELACIONES (fondo naranja)
            for key, data in pagos_totales_otros.items():
                for i, cashflow in enumerate(data['cashflows']):
                    fill_color = PatternFill(start_color="fff3e0", end_color="fff3e0", fill_type="solid")
                    
                    if i == 0:
                        cell1 = ws.cell(row=row, column=1, value=f"{data['order'].subsidiary.serial}-{data['order'].correlative:03d}")
                        cell1.border = border
                        cell1.fill = fill_color
                        
                        cell2 = ws.cell(row=row, column=2, value=data['order'].client.full_name if data['order'].client else '-')
                        cell2.border = border
                        cell2.fill = fill_color
                        
                        cell3 = ws.cell(row=row, column=3, value=1)
                        cell3.border = border
                        cell3.fill = fill_color
                        
                        product_desc = ""
                        if data['order'].orderdetail_set.exists():
                            product_desc = " | ".join([detail.product_name or "Producto Manual" for detail in data['order'].orderdetail_set.all()])
                        else:
                            product_desc = data['order'].observation or "ORDEN DE SERVICIO"
                        cell4 = ws.cell(row=row, column=4, value=product_desc)
                        cell4.border = border
                        cell4.fill = fill_color
                    else:
                        for col in range(1, 5):
                            cell = ws.cell(row=row, column=col, value="")
                            cell.border = border
                            cell.fill = fill_color
                    
                    cell5 = ws.cell(row=row, column=5, value=cashflow.user.first_name or cashflow.user.username or '-')
                    cell5.border = border
                    cell5.fill = fill_color
                    
                    payment_type = "EFECTIVO" if cashflow.way_to_pay == 'E' else ("YAPE" if cashflow.way_to_pay == 'Y' else "DEPÓSITO")
                    cell6 = ws.cell(row=row, column=6, value=payment_type)
                    cell6.border = border
                    cell6.fill = fill_color
                    
                    cell7 = ws.cell(row=row, column=7, value=decimal.Decimal(cashflow.total))
                    cell7.border = border
                    cell7.fill = fill_color
                    cell7.number_format = currency_format
                    
                    if i == 0:
                        cell8 = ws.cell(row=row, column=8, value="CANCELACIÓN")
                        cell8.border = border
                        cell8.fill = PatternFill(start_color="ffe0b2", end_color="ffe0b2", fill_type="solid")
                        cell8.font = Font(bold=True, color="e65100")
                        
                        cell9 = ws.cell(row=row, column=9, value=Decimal(data['order'].total))
                        cell9.border = border
                        cell9.fill = fill_color
                        cell9.number_format = currency_format
                    else:
                        for col in [8, 9]:
                            cell = ws.cell(row=row, column=col, value="")
                            cell.border = border
                            cell.fill = fill_color
                    
                    row += 1
            
            # Totales de ingresos
            row += 1
            ws.cell(row=row, column=8, value="YAPE:").font = Font(bold=True)
            cell_yape = ws.cell(row=row, column=9, value=Decimal(ingresos_yape))
            cell_yape.font = Font(bold=True)
            cell_yape.number_format = currency_format
            row += 1
            ws.cell(row=row, column=8, value="EFECTIVO:").font = Font(bold=True)
            cell_efectivo = ws.cell(row=row, column=9, value=Decimal(ingresos_efectivo))
            cell_efectivo.font = Font(bold=True)
            cell_efectivo.number_format = currency_format
            row += 1
            ws.cell(row=row, column=8, value="TOTAL INGRESOS:").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=8).fill = header_fill_primary
            cell_total_ing = ws.cell(row=row, column=9, value=Decimal(total_ingresos_dia))
            cell_total_ing.font = Font(bold=True, color="FFFFFF")
            cell_total_ing.fill = header_fill_primary
            cell_total_ing.number_format = currency_format
            
            # Sección de SALDOS (solo si hay datos)
            if pagos_fechas_anteriores.exists():
                row += 3  # Espacio entre secciones
                
                # Título de saldos
                ws.cell(row=row, column=1, value="SALDOS")
                ws.cell(row=row, column=1).font = header_font
                ws.cell(row=row, column=1).fill = header_fill_success
                ws.merge_cells(f'A{row}:F{row}')
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                
                # Encabezados de saldos
                row += 1
                payments_headers = ['N° COMPROBANTE', 'FECHA', 'DESCRIPCIÓN', 'USUARIO', 'TIPO PAGO', 'S/TOTAL']
                for col, header in enumerate(payments_headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill_success
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                # Datos de saldos
                row += 1
                for cashflow in pagos_fechas_anteriores:
                    ws.cell(row=row, column=1, value=f"{cashflow.order.subsidiary.serial}-{cashflow.order.correlative:03d}").border = border
                    ws.cell(row=row, column=2, value=cashflow.order.register_date.strftime('%d-%m-%Y')).border = border
                    
                    # Descripción con productos
                    desc = cashflow.description or "PAGO TOTAL"
                    desc += f" (Usuario: {cashflow.order.user.first_name or cashflow.order.user.username or '-'})"
                    if cashflow.order.orderdetail_set.exists():
                        products = " | ".join([detail.product_name or "Producto Manual" for detail in cashflow.order.orderdetail_set.all()])
                        desc += f" - {products}"
                    ws.cell(row=row, column=3, value=desc).border = border
                    
                    ws.cell(row=row, column=4, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                    
                    # Tipo de pago
                    payment_type = ""
                    if cashflow.way_to_pay == 'E':
                        payment_type = "EFECTIVO"
                    elif cashflow.way_to_pay == 'Y':
                        payment_type = "YAPE"
                    elif cashflow.way_to_pay == 'D':
                        payment_type = "DEPÓSITO"
                    ws.cell(row=row, column=5, value=payment_type).border = border
                    cell_saldo_total = ws.cell(row=row, column=6, value=Decimal(cashflow.total))
                    cell_saldo_total.border = border
                    cell_saldo_total.number_format = currency_format
                    row += 1
                
                # Totales de saldos
                row += 1
                ws.cell(row=row, column=5, value="YAPE:").font = Font(bold=True)
                cell_saldo_yape = ws.cell(row=row, column=6, value=Decimal(pagos_anteriores_yape))
                cell_saldo_yape.font = Font(bold=True)
                cell_saldo_yape.number_format = currency_format
                row += 1
                ws.cell(row=row, column=5, value="EFECTIVO:").font = Font(bold=True)
                cell_saldo_efectivo = ws.cell(row=row, column=6, value=Decimal(pagos_anteriores_efectivo))
                cell_saldo_efectivo.font = Font(bold=True)
                cell_saldo_efectivo.number_format = currency_format
                row += 1
                ws.cell(row=row, column=5, value="TOTAL PAGOS ANTERIORES:").font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=5).fill = header_fill_success
                cell_total_pagos_ant = ws.cell(row=row, column=6, value=Decimal(total_pagos_anteriores))
                cell_total_pagos_ant.font = Font(bold=True, color="FFFFFF")
                cell_total_pagos_ant.fill = header_fill_success
                cell_total_pagos_ant.number_format = currency_format
            
            # Sección de EGRESOS (solo si hay datos)
            if expenses_cashflows.exists():
                row += 3  # Espacio entre secciones
                
                # Título de egresos
                ws.cell(row=row, column=1, value="EGRESOS")
                ws.cell(row=row, column=1).font = header_font
                ws.cell(row=row, column=1).fill = header_fill_danger
                ws.merge_cells(f'A{row}:E{row}')
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                
                # Encabezados de egresos
                row += 1
                expenses_headers = ['NRO', 'DESCRIPCIÓN', 'TIPO EGRESO', 'USUARIO', 'MONTO']
                for col, header in enumerate(expenses_headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill_danger
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                # Datos de egresos
                row += 1
                for i, cashflow in enumerate(expenses_cashflows, 1):
                    ws.cell(row=row, column=1, value=i).border = border
                    ws.cell(row=row, column=2, value=cashflow.description or '-').border = border
                    
                    # Tipo de egreso
                    expense_type = ""
                    if cashflow.type_expense == 'V':
                        expense_type = "VARIABLE"
                    elif cashflow.type_expense == 'F':
                        expense_type = "FIJO"
                    elif cashflow.type_expense == 'P':
                        expense_type = "PERSONAL"
                    elif cashflow.type_expense == 'O':
                        expense_type = "OTRO"
                    ws.cell(row=row, column=3, value=expense_type).border = border
                    ws.cell(row=row, column=4, value=cashflow.user.first_name or cashflow.user.username or '-').border = border
                    cell_egreso = ws.cell(row=row, column=5, value=Decimal(cashflow.total))
                    cell_egreso.border = border
                    cell_egreso.number_format = currency_format
                    row += 1
                
                # Total de egresos
                row += 1
                ws.cell(row=row, column=4, value="TOTAL EGRESOS:").font = Font(bold=True, color="FFFFFF")
                ws.cell(row=row, column=4).fill = header_fill_danger
                cell_total_egreso = ws.cell(row=row, column=5, value=Decimal(total_expenses_amount))
                cell_total_egreso.font = Font(bold=True, color="FFFFFF")
                cell_total_egreso.fill = header_fill_danger
                cell_total_egreso.number_format = currency_format
            
            # Sección de RESUMENES
            row += 3  # Espacio entre secciones
            
            # Título de resúmenes
            ws.cell(row=row, column=1, value="RESUMENES")
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=1).fill = header_fill_success
            ws.merge_cells(f'A{row}:B{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Datos de resúmenes
            row += 1
            ws.cell(row=row, column=1, value="INGRESOS").font = Font(bold=True, size=12, color="007bff")
            row += 1
            ws.cell(row=row, column=1, value="APERTURA DE CAJA:").font = Font(bold=True)
            cell_res_apertura = ws.cell(row=row, column=2, value=Decimal(total_apertura))
            cell_res_apertura.font = Font(bold=True)
            cell_res_apertura.number_format = currency_format
            row += 1
            ws.cell(row=row, column=1, value="INGRESOS DEL DÍA:").font = Font(bold=True)
            cell_res_ing_dia = ws.cell(row=row, column=2, value=Decimal(total_ingresos_dia))
            cell_res_ing_dia.font = Font(bold=True)
            cell_res_ing_dia.number_format = currency_format
            row += 1
            ws.cell(row=row, column=1, value="PAGOS ANTERIORES:").font = Font(bold=True)
            cell_res_pagos_ant = ws.cell(row=row, column=2, value=Decimal(total_pagos_anteriores))
            cell_res_pagos_ant.font = Font(bold=True)
            cell_res_pagos_ant.number_format = currency_format
            row += 1
            ws.cell(row=row, column=1, value="SUBTOTAL INGRESOS:").font = Font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
            cell_res_subtotal = ws.cell(row=row, column=2, value=Decimal(total_apertura + total_ingresos_dia + total_pagos_anteriores))
            cell_res_subtotal.font = Font(bold=True)
            cell_res_subtotal.fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
            cell_res_subtotal.number_format = currency_format
            
            row += 2
            ws.cell(row=row, column=1, value="EGRESOS").font = Font(bold=True, size=12, color="dc3545")
            row += 1
            ws.cell(row=row, column=1, value="TOTAL EGRESOS:").font = Font(bold=True)
            cell_res_egresos = ws.cell(row=row, column=2, value=Decimal(total_expenses_amount))
            cell_res_egresos.font = Font(bold=True)
            cell_res_egresos.number_format = currency_format
            
            row += 3
            ws.cell(row=row, column=1, value="TOTAL EFECTIVO:").font = Font(bold=True, size=11, color="28a745")
            cell_res_efectivo = ws.cell(row=row, column=2, value=Decimal(total_efectivo))
            cell_res_efectivo.font = Font(bold=True, size=11, color="28a745")
            cell_res_efectivo.number_format = currency_format
            row += 1
            ws.cell(row=row, column=1, value="TOTAL YAPE:").font = Font(bold=True, size=11, color="17a2b8")
            cell_res_yape = ws.cell(row=row, column=2, value=Decimal(total_yape))
            cell_res_yape.font = Font(bold=True, size=11, color="17a2b8")
            cell_res_yape.number_format = currency_format
            row += 1
            ws.cell(row=row, column=1, value="APERTURA CAJA:").font = Font(bold=True, size=11, color="007bff")
            cell_res_apertura2 = ws.cell(row=row, column=2, value=Decimal(total_apertura))
            cell_res_apertura2.font = Font(bold=True, size=11, color="007bff")
            cell_res_apertura2.number_format = currency_format
            row += 1
            ws.cell(row=row, column=1, value="TOTAL EGRESOS:").font = Font(bold=True, size=11, color="dc3545")
            cell_res_egresos2 = ws.cell(row=row, column=2, value=Decimal(total_expenses_amount))
            cell_res_egresos2.font = Font(bold=True, size=11, color="dc3545")
            cell_res_egresos2.number_format = currency_format
            row += 1
            ws.cell(row=row, column=1, value="TOTAL FINAL:").font = Font(bold=True, size=12, color="ffc107")
            ws.cell(row=row, column=1).fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            cell_res_final = ws.cell(row=row, column=2, value=Decimal(total_general - total_expenses_amount))
            cell_res_final.font = Font(bold=True, size=12, color="ffc107")
            cell_res_final.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            cell_res_final.number_format = currency_format
            
            # Ajustar ancho de columnas
            for col in range(1, 10):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Guardar archivo
            filename = f"reporte_ventas_usuario_{report_date}.xlsx"
            file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            wb.save(file_path)
            
            # URL del archivo
            file_url = f"{settings.MEDIA_URL}reports/{filename}"
            
            return JsonResponse({
                'success': True,
                'message': 'Reporte Excel generado exitosamente',
                'file_url': file_url,
                'filename': filename
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al generar el Excel: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)


# =============================================================================
# REPORTES PARA ACADEMIA DE BASKETBALL
# =============================================================================

@csrf_exempt
def export_enrollments_report_excel(request):
    """Exportar reporte de inscripciones del mes a Excel"""
    if request.method == 'POST':
        try:
            from ..students.models import Enrollment, Student, Cycle, TrainingSchedule, TimeSlot
            
            # Obtener datos del reporte
            report_month = request.POST.get('report_month')
            subsidiary_id = request.POST.get('subsidiary')
            
            if not report_month:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar un mes'
                }, status=HTTPStatus.BAD_REQUEST)
            
            # Convertir mes a rango de fechas
            year, month = report_month.split('-')
            start_date = datetime(int(year), int(month), 1).date()
            if int(month) == 12:
                end_date = datetime(int(year) + 1, 1, 1).date()
            else:
                end_date = datetime(int(year), int(month) + 1, 1).date()
            
            # Filtrar inscripciones
            enrollments_filter = {}
            subsidiary_obj = None
            
            if subsidiary_id and subsidiary_id != '0':
                subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary_id))
                enrollments_filter = {'subsidiary_id': subsidiary_id}
            
            enrollments = Enrollment.objects.filter(
                enrollment_date__gte=start_date,
                enrollment_date__lt=end_date,
                **enrollments_filter
            ).select_related(
                'student', 'cycle', 'schedule', 'time_slot', 'time_slot__teacher', 'subsidiary', 'user'
            ).order_by('enrollment_date', 'student__full_name')
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Inscripciones {report_month}"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
            title_font = Font(bold=True, size=14, color="007bff")
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            currency_format = '_("S/"* #,##0.00_);_("S/"* (#,##0.00);_("S/"* "-"??_);_(@_)'
            
            # Título
            ws.merge_cells('A1:L1')
            subsidiary_name = subsidiary_obj.name.upper() if subsidiary_obj else 'TODAS LAS SEDES'
            month_name = start_date.strftime('%B %Y').upper()
            ws['A1'] = f"REPORTE DE INSCRIPCIONES - {subsidiary_name} - {month_name}"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Encabezados
            headers = [
                'N°', 'Fecha', 'Deportista', 'DNI', 'Ciclo', 'Horario', 'Turno', 
                'Profesor', 'Sede', 'Total S/.', 'Adelanto S/.', 'Faltante S/.'
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Datos
            row = 4
            for idx, enrollment in enumerate(enrollments, 1):
                ws.cell(row=row, column=1, value=idx).border = border
                ws.cell(row=row, column=2, value=enrollment.enrollment_date.strftime('%d/%m/%Y')).border = border
                ws.cell(row=row, column=3, value=enrollment.student.full_name or '-').border = border
                ws.cell(row=row, column=4, value=enrollment.student.document_number or '-').border = border
                ws.cell(row=row, column=5, value=enrollment.cycle.name).border = border
                ws.cell(row=row, column=6, value=enrollment.schedule.name).border = border
                
                time_slot_str = f"{enrollment.time_slot.start_time.strftime('%H:%M')} - {enrollment.time_slot.end_time.strftime('%H:%M')}"
                if enrollment.time_slot.age_description:
                    time_slot_str += f" ({enrollment.time_slot.age_description})"
                ws.cell(row=row, column=7, value=time_slot_str).border = border
                
                teacher_name = '-'
                if enrollment.time_slot.teacher:
                    teacher_name = f"{enrollment.time_slot.teacher.first_name or ''} {enrollment.time_slot.teacher.last_name or ''}".strip() or enrollment.time_slot.teacher.username
                ws.cell(row=row, column=8, value=teacher_name).border = border
                
                ws.cell(row=row, column=9, value=enrollment.subsidiary.name if enrollment.subsidiary else '-').border = border
                
                cell_total = ws.cell(row=row, column=10, value=Decimal(enrollment.price))
                cell_total.border = border
                cell_total.number_format = currency_format
                
                cell_advance = ws.cell(row=row, column=11, value=Decimal(enrollment.advance))
                cell_advance.border = border
                cell_advance.number_format = currency_format
                
                cell_remaining = ws.cell(row=row, column=12, value=Decimal(enrollment.remaining))
                cell_remaining.border = border
                cell_remaining.number_format = currency_format
                
                row += 1
            
            # Totales
            row += 1
            ws.merge_cells(f'A{row}:I{row}')
            ws.cell(row=row, column=1, value="TOTALES:").font = Font(bold=True, size=12)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
            
            total_price = enrollments.aggregate(Sum('price'))['price__sum'] or Decimal('0')
            total_advance = enrollments.aggregate(Sum('advance'))['advance__sum'] or Decimal('0')
            total_remaining = enrollments.aggregate(Sum('remaining'))['remaining__sum'] or Decimal('0')
            
            cell_total_price = ws.cell(row=row, column=10, value=Decimal(total_price))
            cell_total_price.font = Font(bold=True, size=12)
            cell_total_price.fill = header_fill
            cell_total_price.font = Font(bold=True, color="FFFFFF", size=12)
            cell_total_price.number_format = currency_format
            cell_total_price.border = border
            
            cell_total_advance = ws.cell(row=row, column=11, value=Decimal(total_advance))
            cell_total_advance.font = Font(bold=True, size=12)
            cell_total_advance.fill = header_fill
            cell_total_advance.font = Font(bold=True, color="FFFFFF", size=12)
            cell_total_advance.number_format = currency_format
            cell_total_advance.border = border
            
            cell_total_remaining = ws.cell(row=row, column=12, value=Decimal(total_remaining))
            cell_total_remaining.font = Font(bold=True, size=12)
            cell_total_remaining.fill = header_fill
            cell_total_remaining.font = Font(bold=True, color="FFFFFF", size=12)
            cell_total_remaining.number_format = currency_format
            cell_total_remaining.border = border
            
            # Ajustar ancho de columnas
            column_widths = [6, 12, 30, 12, 20, 15, 20, 20, 15, 12, 12, 12]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Guardar archivo
            filename = f"reporte_inscripciones_{report_month}.xlsx"
            filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            wb.save(filepath)
            
            file_url = f"{settings.MEDIA_URL}reports/{filename}"
            
            return JsonResponse({
                'success': True,
                'message': 'Reporte exportado exitosamente',
                'file_url': file_url,
                'filename': filename
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al exportar el reporte: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)


@csrf_exempt
def export_students_report_excel(request):
    """Exportar reporte de deportistas activos a Excel"""
    if request.method == 'POST':
        try:
            from ..students.models import Student, Enrollment
            
            # Obtener datos del reporte
            subsidiary_id = request.POST.get('subsidiary', '0')
            
            # Filtrar estudiantes
            students_filter = {'is_active': True}
            subsidiary_obj = None
            
            if subsidiary_id and subsidiary_id != '0':
                subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary_id))
                students_filter['subsidiary_id'] = subsidiary_id
            
            students = Student.objects.filter(**students_filter).order_by('full_name', 'surname')
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Deportistas Activos"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            title_font = Font(bold=True, size=14, color="28a745")
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Título
            ws.merge_cells('A1:J1')
            subsidiary_name = subsidiary_obj.name.upper() if subsidiary_obj else 'TODAS LAS SEDES'
            ws['A1'] = f"REPORTE DE DEPORTISTAS ACTIVOS - {subsidiary_name}"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Encabezados
            headers = [
                'N°', 'Nombres y Apellidos', 'DNI', 'Fecha Nacimiento', 'Edad', 
                'Teléfono', 'Padre/Madre/Apoderado', 'Teléfono Apoderado', 'Sede', 'Fecha Registro'
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Datos
            row = 4
            for idx, student in enumerate(students, 1):
                ws.cell(row=row, column=1, value=idx).border = border
                ws.cell(row=row, column=2, value=student.full_name or '-').border = border
                ws.cell(row=row, column=3, value=student.document_number or '-').border = border
                ws.cell(row=row, column=4, value=student.birth_date.strftime('%d/%m/%Y') if student.birth_date else '-').border = border
                ws.cell(row=row, column=5, value=student.current_age or '-').border = border
                ws.cell(row=row, column=6, value=student.phone or '-').border = border
                
                parent_name = f"{student.parent_first_name or ''} {student.parent_surname or ''}".strip() or '-'
                ws.cell(row=row, column=7, value=parent_name).border = border
                ws.cell(row=row, column=8, value=student.parent_phone or '-').border = border
                ws.cell(row=row, column=9, value=student.subsidiary.name if student.subsidiary else '-').border = border
                ws.cell(row=row, column=10, value=student.creation_date.strftime('%d/%m/%Y %H:%M') if student.creation_date else '-').border = border
                
                row += 1
            
            # Total
            row += 1
            ws.merge_cells(f'A{row}:I{row}')
            ws.cell(row=row, column=1, value=f"TOTAL DE DEPORTISTAS: {students.count()}").font = Font(bold=True, size=12)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
            
            # Ajustar ancho de columnas
            column_widths = [6, 35, 12, 15, 8, 15, 30, 15, 20, 18]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Guardar archivo
            filename = f"reporte_deportistas_activos_{datetime.now().strftime('%Y%m%d')}.xlsx"
            filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            wb.save(filepath)
            
            file_url = f"{settings.MEDIA_URL}reports/{filename}"
            
            return JsonResponse({
                'success': True,
                'message': 'Reporte exportado exitosamente',
                'file_url': file_url,
                'filename': filename
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al exportar el reporte: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)


@csrf_exempt
def export_pending_payments_report_excel(request):
    """Exportar reporte de pagos pendientes a Excel"""
    if request.method == 'POST':
        try:
            from ..students.models import Enrollment
            
            # Obtener datos del reporte
            subsidiary_id = request.POST.get('subsidiary', '0')
            
            # Filtrar inscripciones con pagos pendientes
            enrollments_filter = {'remaining__gt': 0}
            subsidiary_obj = None
            
            if subsidiary_id and subsidiary_id != '0':
                subsidiary_obj = Subsidiary.objects.get(id=int(subsidiary_id))
                enrollments_filter['subsidiary_id'] = subsidiary_id
            
            enrollments = Enrollment.objects.filter(**enrollments_filter).select_related(
                'student', 'cycle', 'schedule', 'time_slot', 'subsidiary'
            ).order_by('enrollment_date')
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pagos Pendientes"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            title_font = Font(bold=True, size=14, color="dc3545")
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            currency_format = '_("S/"* #,##0.00_);_("S/"* (#,##0.00);_("S/"* "-"??_);_(@_)'
            
            # Título
            ws.merge_cells('A1:J1')
            subsidiary_name = subsidiary_obj.name.upper() if subsidiary_obj else 'TODAS LAS SEDES'
            ws['A1'] = f"REPORTE DE PAGOS PENDIENTES - {subsidiary_name}"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Encabezados
            headers = [
                'N°', 'Fecha Inscripción', 'Deportista', 'DNI', 'Ciclo', 
                'Total S/.', 'Adelanto S/.', 'Faltante S/.', 'Sede', 'Días Pendiente'
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Datos
            row = 4
            today = datetime.now().date()
            for idx, enrollment in enumerate(enrollments, 1):
                days_pending = (today - enrollment.enrollment_date).days
                
                ws.cell(row=row, column=1, value=idx).border = border
                ws.cell(row=row, column=2, value=enrollment.enrollment_date.strftime('%d/%m/%Y')).border = border
                ws.cell(row=row, column=3, value=enrollment.student.full_name or '-').border = border
                ws.cell(row=row, column=4, value=enrollment.student.document_number or '-').border = border
                ws.cell(row=row, column=5, value=enrollment.cycle.name).border = border
                
                cell_total = ws.cell(row=row, column=6, value=Decimal(enrollment.price))
                cell_total.border = border
                cell_total.number_format = currency_format
                
                cell_advance = ws.cell(row=row, column=7, value=Decimal(enrollment.advance))
                cell_advance.border = border
                cell_advance.number_format = currency_format
                
                cell_remaining = ws.cell(row=row, column=8, value=Decimal(enrollment.remaining))
                cell_remaining.border = border
                cell_remaining.number_format = currency_format
                if enrollment.remaining > 0:
                    cell_remaining.font = Font(bold=True, color="dc3545")
                
                ws.cell(row=row, column=9, value=enrollment.subsidiary.name if enrollment.subsidiary else '-').border = border
                ws.cell(row=row, column=10, value=days_pending).border = border
                
                row += 1
            
            # Totales
            row += 1
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1, value="TOTALES:").font = Font(bold=True, size=12)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
            
            total_price = enrollments.aggregate(Sum('price'))['price__sum'] or Decimal('0')
            total_advance = enrollments.aggregate(Sum('advance'))['advance__sum'] or Decimal('0')
            total_remaining = enrollments.aggregate(Sum('remaining'))['remaining__sum'] or Decimal('0')
            
            cell_total_price = ws.cell(row=row, column=6, value=Decimal(total_price))
            cell_total_price.font = Font(bold=True, color="FFFFFF", size=12)
            cell_total_price.fill = header_fill
            cell_total_price.number_format = currency_format
            cell_total_price.border = border
            
            cell_total_advance = ws.cell(row=row, column=7, value=Decimal(total_advance))
            cell_total_advance.font = Font(bold=True, color="FFFFFF", size=12)
            cell_total_advance.fill = header_fill
            cell_total_advance.number_format = currency_format
            cell_total_advance.border = border
            
            cell_total_remaining = ws.cell(row=row, column=8, value=Decimal(total_remaining))
            cell_total_remaining.font = Font(bold=True, color="FFFFFF", size=12)
            cell_total_remaining.fill = header_fill
            cell_total_remaining.number_format = currency_format
            cell_total_remaining.border = border
            
            # Ajustar ancho de columnas
            column_widths = [6, 15, 30, 12, 20, 12, 12, 12, 15, 12]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Guardar archivo
            filename = f"reporte_pagos_pendientes_{datetime.now().strftime('%Y%m%d')}.xlsx"
            filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            wb.save(filepath)
            
            file_url = f"{settings.MEDIA_URL}reports/{filename}"
            
            return JsonResponse({
                'success': True,
                'message': 'Reporte exportado exitosamente',
                'file_url': file_url,
                'filename': filename
            }, status=HTTPStatus.OK)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al exportar el reporte: {str(e)}'
            }, status=HTTPStatus.INTERNAL_SERVER_ERROR)
    
    return JsonResponse({'message': 'Error de petición.'}, status=HTTPStatus.BAD_REQUEST)